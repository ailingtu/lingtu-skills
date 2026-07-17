"""argparse 解析与子命令实现。"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Any

def _shared_scripts_dir() -> Path:
    current = Path(__file__).resolve()
    for parent in current.parents:
        candidate = parent / "shared" / "scripts"
        if candidate.exists():
            return candidate
    raise RuntimeError("未找到 shared/scripts 目录。请确认 skill 安装完整。")


sys.path.insert(0, str(_shared_scripts_dir()))
# lingtu_auth is used via .api.require_api_key

from .api import (
    creator_can_publish_photo,
    create_post,
    extract_permissions,
    is_tiktok_shop_auth_source,
    list_creator_accounts,
    list_shop_products,
    require_api_key,
    resolve_creator_batch,
    upload_file,
)
from .config import (
    DEFAULT_DESKTOP,
    DEFAULT_PHOTO_POST_TYPE,
    IMAGE_EXTENSIONS,
    PHOTO_MAX_IMAGES,
    PHOTO_SHOPPABLE_PERMISSION,
    PLATFORM_LABELS,
    PRODUCT_TITLE_MAX_LENGTH,
    POST_TITLE_MAX_LENGTH,
    PUBLISH_RECORDS_URL,
    REGION_TIMEZONE_MAP,
    SUPPORTED_MEDIA_TYPES,
    SUPPORTED_PLATFORMS,
)
from .image_utils import format_photo_constraints_help, validate_photo_files
from .excel_utils import (
    COLUMN_LABELS,
    COLUMN_LABEL_TO_KEY,
    CSV_ALL_COLUMNS,
    generate_csv_template,
    generate_excel_template,
    parse_date_for_filename,
    parse_datetime_to_epoch_ms,
    parse_excel_or_csv,
    parse_media_type,
    parse_timezone,
    read_csv_columns,
    has_unsupported_plain_text,
    sanitize_post_title,
    sanitize_product_title,
    split_media_filenames,
    write_csv_schedule,
)
from .report import format_creators, format_products, format_publish_results
from .scheduler import auto_assign_schedule, build_schedule_rows, compute_schedule_times


def print_json(data: Any) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2))


def _format_preview_table(rows_data: list[dict[str, str]], times: list[str], tz: str, dates: list[str]) -> str:
    """将排期数据格式化为聊天预览文本。"""
    from collections import defaultdict

    is_shop = rows_data[0].get("platform") == "tiktok_shop"
    is_photo = parse_media_type(rows_data[0].get("media_type")) == "photo"

    # 按日期+达人分组统计
    by_day_creator: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in rows_data:
        parts = r["scheduled_at"].split()
        d = parts[0]
        c = r["creator_username"]
        by_day_creator[d][c] += 1

    if is_photo:
        type_label = "带货图文"
    elif is_shop:
        type_label = "带货视频"
    else:
        type_label = "养号视频"

    lines = ["📋 排期预览：", ""]
    lines.append(f"类型：{type_label}")
    lines.append(f"时区：{tz}")
    lines.append(f"默认基础时间段：{' / '.join(times)}（达人间按分钟错峰）")
    if is_shop:
        lines.append(f"产品ID：{rows_data[0].get('product_id') or '无'}")
    lines.append("")

    # 按日期汇总
    for d in dates:
        dc = by_day_creator[d]
        total = sum(dc.values())
        creators_str = "、".join(f"{c}({n}条)" for c, n in dc.items())
        lines.append(f"{d}：共 {total} 条 — {creators_str}")

    lines.append("")
    all_creators = set(r["creator_username"] for r in rows_data)
    unit = "条图文" if is_photo else "条视频"
    lines.append(f"共 {len(dates)} 天 · {len(rows_data)} {unit} · {len(all_creators)} 个达人")
    lines.append("")
    lines.append("确认后生成 CSV，你只需补充：")
    if is_photo:
        lines.append("  · 购物车标题 · 视频文案内容 · 图片文件名（多图用逗号分隔）")
        lines.append(f"  · {format_photo_constraints_help()}")
        lines.append("  · （可选）音乐ID / 音乐标题 / 音乐作者 / 音乐时长")
    elif is_shop:
        lines.append("  · 购物车标题 · 视频文案内容 · 视频文件名")
    else:
        lines.append("  · 视频文案内容 · 视频文件名")

    return "\n".join(lines)


def _csv_columns_for_platform(platform: str, media_type: str = "video") -> tuple[str, ...]:
    is_photo = parse_media_type(media_type) == "photo"
    if platform == "tiktok":
        return (
            "creator_username", "platform", "media_type",
            "title", "timezone",
            "scheduled_at", "video_file",
        )
    if is_photo:
        return (
            "creator_username", "platform", "media_type",
            "product_id", "product_title", "product_source",
            "title", "timezone", "scheduled_at",
            "image_files",
            "music_id", "music_title", "music_author", "music_duration",
        )
    # 带货视频也带「媒体类」列，默认填 video；用户可改成 photo
    return (
        "creator_username", "platform", "media_type",
        "product_id", "product_title",
        "product_source", "title",
        "timezone", "scheduled_at",
        "video_file",
    )

def _timezone_from_creator_info(creator_info: dict[str, Any]) -> str:
    for key in (
        "targetRegion",
        "targetMarket",
        "marketRegion",
        "shopRegion",
        "selectionRegion",
        "oauthRegion",
        "registerRegion",
    ):
        region = str(creator_info.get(key) or "").strip()
        if not region:
            continue
        normalized = region.upper().replace("-", "_")
        if normalized in REGION_TIMEZONE_MAP:
            return REGION_TIMEZONE_MAP[normalized]
    return "America/Los_Angeles"


def _timezone_map_for_creators(
    creators: list[str],
    creator_info_by_username: dict[str, dict[str, Any]] | None,
    timezone_override: str,
    region_hint: str = "",
) -> dict[str, str]:
    if timezone_override:
        return {creator.lower(): timezone_override for creator in creators}
    if region_hint:
        region_tz = _timezone_from_creator_info({"targetRegion": region_hint})
        return {creator.lower(): region_tz for creator in creators}
    creator_info_by_username = creator_info_by_username or {}
    return {
        creator.lower(): _timezone_from_creator_info(creator_info_by_username.get(creator.lower(), {}))
        for creator in creators
    }


def _normalize_region(raw: str | None) -> str:
    return (raw or "").strip().upper().replace("-", "_")


def _parse_daily_counts(raw: str, start_year: int) -> dict[str, int]:
    """Parse YYYY-MM-DD=N or MM-DD=N entries into a date -> count map."""
    if not raw.strip():
        return {}

    import re

    counts: dict[str, int] = {}
    for item in raw.split(","):
        part = item.strip()
        if not part:
            continue
        if "=" not in part:
            raise SystemExit(f"--daily-counts 格式错误：{part}（需要 YYYY-MM-DD=条数）")
        date_part, count_part = [value.strip() for value in part.split("=", 1)]
        if not count_part.isdigit() or int(count_part) < 1:
            raise SystemExit(f"--daily-counts 条数错误：{part}（条数必须为正整数）")
        if re.match(r"^\d{4}-\d{2}-\d{2}$", date_part):
            date_key = parse_date_for_filename(date_part)
        elif re.match(r"^\d{1,2}-\d{1,2}$", date_part):
            month, day = [int(value) for value in date_part.split("-", 1)]
            date_key = parse_date_for_filename(f"{start_year:04d}-{month:02d}-{day:02d}")
        else:
            raise SystemExit(f"--daily-counts 日期格式错误：{date_part}（需要 YYYY-MM-DD 或 MM-DD）")
        counts[date_key] = int(count_part)
    return counts


# ── gen-csv ──────────────────────────────────────────────────

def command_gen_csv(args: argparse.Namespace) -> None:
    platform = args.platform
    media_type = parse_media_type(getattr(args, "media_type", None) or "video")
    if media_type == "photo" and platform != "tiktok_shop":
        raise SystemExit("带货图文 (media-type=photo) 仅支持 --platform tiktok_shop")
    date_str = parse_date_for_filename(args.date)
    tz_iana = parse_timezone(args.timezone) if args.timezone else ""
    region_hint = _normalize_region(getattr(args, "region", ""))
    count = max(1, args.count)
    days = max(1, args.days)

    from datetime import datetime, timedelta

    # 计算多天日期列表
    start_dt = datetime.strptime(date_str, "%Y-%m-%d")
    count_by_date = _parse_daily_counts(getattr(args, "daily_counts", "") or "", start_dt.year)
    if count_by_date:
        dates = sorted(count_by_date)
        preview_count = max(count_by_date.values())
    else:
        dates = [(start_dt + timedelta(days=d)).strftime("%Y-%m-%d") for d in range(days)]
        preview_count = count
    times = compute_schedule_times(preview_count)

    product_id = args.product_id or ""
    if platform == "tiktok":
        product_id = ""

    # 获取达人列表（dry-run 先不调 API，用用户输入直接预览）
    raw_creators = [u.strip() for u in args.creators.split(",") if u.strip()] if args.creators else []

    if args.dry_run:
        creator_list = raw_creators if raw_creators else ["（全部已授权达人）"]
        preview_tz = tz_iana or _timezone_from_creator_info({"targetRegion": region_hint})
        rows_data = build_schedule_rows(
            dates=dates,
            creators=creator_list,
            platform=platform,
            product_id=product_id,
            timezone_by_creator={creator.lower(): preview_tz for creator in creator_list},
            count=count,
            count_by_date=count_by_date,
            media_type=media_type,
        )
        print(_format_preview_table(rows_data, times, preview_tz, dates))
        return

    # ── 正式生成：调 API 验证达人 ──
    require_api_key()

    photo_filter = True if media_type == "photo" else None

    if raw_creators:
        creator_list = raw_creators
    else:
        selection_region = region_hint if platform == "tiktok_shop" and region_hint else None
        result = list_creator_accounts(
            platform=platform,
            selection_region=selection_region,
            has_photo_permission=photo_filter,
        )
        data = result.get("data") or {}
        items = data.get("list") or []
        # 服务端 hasPhotoPermission 为主；客户端再兜底一次
        if media_type == "photo":
            items = [item for item in items if _item_can_publish_photo(item)]
        creator_list = [item.get("username", "") for item in items if item.get("username")]
        if not creator_list:
            region_msg = f"、地区 {region_hint}" if selection_region else ""
            if media_type == "photo":
                raise SystemExit(
                    f"该平台（{PLATFORM_LABELS.get(platform, platform)}{region_msg}）下没有可发带货图文的达人。"
                    f"需 TikTok Shop 授权且 permissions 含 {PHOTO_SHOPPABLE_PERMISSION}（可传 hasPhotoPermission=true 筛选）。"
                    "请先打开 https://app.ailingtu.com/video-post 授权后再重试。"
                )
            raise SystemExit(
                f"该平台（{PLATFORM_LABELS.get(platform, platform)}{region_msg}）下没有已授权的达人。"
                "请先打开 https://app.ailingtu.com/video-post 授权达人后再重试。"
            )

    # 验证指定的达人是否存在
    creator_info_by_username: dict[str, dict[str, Any]] = {}
    if raw_creators:
        found, not_found = resolve_creator_batch(
            creator_list,
            platform=platform,
            has_photo_permission=photo_filter,
        )
        if not_found:
            if media_type == "photo":
                print(
                    f"⚠ 以下达人未找到、未授权，或无图文权限（hasPhotoPermission）：{', '.join(not_found)}",
                    file=sys.stderr,
                )
            else:
                print(f"⚠ 以下达人未找到或未授权：{', '.join(not_found)}", file=sys.stderr)
        if media_type == "photo" and found:
            photo_ok: dict[str, dict[str, Any]] = {}
            photo_blocked: list[str] = []
            for uname, info in found.items():
                ok, reason = creator_can_publish_photo(info)
                if ok:
                    photo_ok[uname] = info
                else:
                    photo_blocked.append(reason or uname)
            if photo_blocked:
                print(
                    "⚠ 以下达人不可发带货图文（需 TikTok Shop + "
                    f"{PHOTO_SHOPPABLE_PERMISSION}）：\n  - "
                    + "\n  - ".join(photo_blocked),
                    file=sys.stderr,
                )
            found = photo_ok
        if not found:
            raise SystemExit(
                "所有指定的达人均未找到、未授权，或不可发带货图文。"
                f"{'图文需 TikTok Shop 授权且 permissions 含 ' + PHOTO_SHOPPABLE_PERMISSION + '。' if media_type == 'photo' else ''}"
                "请先打开 https://app.ailingtu.com/video-post 授权达人后再重试。"
            )
        creator_list = list(found.keys())
        creator_info_by_username = found
    else:
        creator_info_by_username = {
            str(item.get("username", "")).lower(): {
                "targetRegion": item.get("targetRegion") or "",
                "targetMarket": item.get("targetMarket") or "",
                "marketRegion": item.get("marketRegion") or "",
                "shopRegion": item.get("shopRegion") or "",
                "selectionRegion": item.get("selectionRegion") or "",
                "oauthRegion": item.get("oauthRegion") or "",
                "registerRegion": item.get("registerRegion") or "",
                "authSource": item.get("authSource") or "",
                "permissions": extract_permissions(item),
            }
            for item in items
            if item.get("username")
        }

    timezone_by_creator = _timezone_map_for_creators(creator_list, creator_info_by_username, tz_iana, region_hint)

    # 构建所有行
    rows_data = build_schedule_rows(
        dates=dates,
        creators=creator_list,
        platform=platform,
        product_id=product_id,
        timezone_by_creator=timezone_by_creator,
        count=count,
        count_by_date=count_by_date,
        media_type=media_type,
    )

    # 生成文件夹和 CSV
    first_date = dates[0]
    last_date = dates[-1]
    folder_prefix = "图文发布" if media_type == "photo" else "视频发布"
    if len(dates) == 1:
        folder_name = f"{folder_prefix}_{first_date}"
    else:
        folder_name = f"{folder_prefix}_{first_date}_to_{last_date}"
    output_dir = args.output_dir or str(DEFAULT_DESKTOP / folder_name)
    output_dir = str(Path(output_dir).expanduser())
    csv_path = str(Path(output_dir) / "schedule.csv")

    generated = generate_csv_template(
        output_path=csv_path,
        rows_data=rows_data,
        columns=_csv_columns_for_platform(platform, media_type),
    )

    output = {
        "folder": output_dir,
        "csv": generated,
        "schedule": generated,
        "platform": platform,
        "media_type": media_type,
        "date": first_date,
        "days": len(dates),
        "dates": dates,
        "timezone": tz_iana or "auto",
        "region": region_hint,
        "region_filter_applied": platform == "tiktok_shop" and bool(region_hint) and not raw_creators,
        "timezones_by_creator": timezone_by_creator,
        "creators": creator_list,
        "creators_count": len(creator_list),
        "videos_per_creator_per_day": count,
        "daily_counts": count_by_date,
        "total_rows": len(rows_data),
    }

    if args.format == "json":
        print_json(output)
        return

    media_label = "图片" if media_type == "photo" else "视频"
    print(f"已在桌面创建排期文件夹：{output_dir}")
    print(f"  ├── schedule.csv  ← 排期表（已预填达人、时间）")
    print(f"  └── （请将{media_label}文件拖入此文件夹）")
    print()
    print(f"达人数量：{len(creator_list)}")
    print(f"媒体类型：{'带货图文' if media_type == 'photo' else '视频'}")
    print(f"日期范围：{first_date} ~ {last_date}（{len(dates)} 天）")
    if count_by_date:
        print("每达人每日：" + "，".join(f"{d} {n} 条" for d, n in count_by_date.items()))
    else:
        print(f"每达人每日：{count} 条")
    if tz_iana:
        print(f"发布时间：{', '.join(times)}（{tz_iana}，达人间已错峰）")
    else:
        unique_tz = sorted(set(timezone_by_creator.values()))
        print(f"发布时间：{', '.join(times)}（按达人区域自动时区：{', '.join(unique_tz)}；达人间已错峰）")
    if region_hint:
        if platform == "tiktok_shop" and not raw_creators:
            print(f"达人筛选：带货达人列表已按地区 {region_hint} 筛选")
        elif platform == "tiktok":
            print(f"达人筛选：普通/养号达人列表不支持按国家筛选，地区 {region_hint} 仅用于默认时区")
    print(f"总排期行数：{len(rows_data)}")
    print()
    print("下一步：")
    if media_type == "photo":
        print("  1. 打开 schedule.csv，填写 product_title、title 和 image_files（多图逗号分隔）")
        print(f"  2. 图片约束：{format_photo_constraints_help()}")
        print("  3. （可选）填写 music_id / music_title / music_author / music_duration")
        print("  4. 将所有图片文件复制到文件夹内")
        print(f"  5. 运行：lingtu_video_publish.py publish --folder {output_dir} --confirm")
    elif platform == "tiktok":
        print("  1. 打开 schedule.csv，填写 title 和 video_file（填文件名即可）")
        print("  2. 将所有视频文件复制到文件夹内")
        print(f"  3. 运行：lingtu_video_publish.py publish --folder {output_dir} --confirm")
    else:
        print("  1. 打开 schedule.csv，填写 product_title、title 和 video_file（填文件名即可）")
        print("  2. 将所有视频文件复制到文件夹内")
        print(f"  3. 运行：lingtu_video_publish.py publish --folder {output_dir} --confirm")


# ── creators ─────────────────────────────────────────────────

def command_creators(args: argparse.Namespace) -> None:
    platform = args.platform or None
    region = _normalize_region(getattr(args, "region", ""))
    selection_region = region if platform == "tiktok_shop" and region else None
    has_photo = getattr(args, "has_photo_permission", False) or None
    if has_photo:
        has_photo = True
    else:
        has_photo = None
    result = list_creator_accounts(
        platform=platform,
        page_size=args.page_size,
        selection_region=selection_region,
        has_photo_permission=has_photo,
    )
    data = result.get("data") or {}
    items = data.get("list") or []

    if args.username:
        keyword = args.username.strip().lower()
        items = [i for i in items if keyword in (i.get("username") or "").lower()]

    # 标注图文能力，便于筛选（服务端 hasPhotoPermission 筛过则均为 true）
    for item in items:
        can_photo, _ = creator_can_publish_photo({
            "username": item.get("username") or "",
            "authSource": item.get("authSource") or "",
            "permissions": extract_permissions(item),
        })
        item["canPublishPhoto"] = can_photo if not has_photo else True

    if args.format == "json":
        print_json({"creators": items, "total": len(items)})
        return

    print(format_creators(items))


# ── publish ──────────────────────────────────────────────────

def command_publish(args: argparse.Namespace) -> None:
    folder = Path(args.folder).expanduser()
    if not folder.exists():
        raise SystemExit(f"文件夹不存在：{folder}")
    if not folder.is_dir():
        raise SystemExit(f"路径不是文件夹：{folder}")

    # 查找排期文件
    csv_file = folder / "schedule.csv"
    excel_file = folder / "schedule.xlsx"
    if csv_file.exists():
        schedule_path = str(csv_file)
    elif excel_file.exists():
        schedule_path = str(excel_file)
    else:
        raise SystemExit(f"文件夹内未找到 schedule.xlsx 或 schedule.csv：{folder}")

    rows = parse_excel_or_csv(schedule_path)
    if not rows:
        raise SystemExit("排期表为空。")

    # 校验每行
    validation_errors: list[dict[str, Any]] = []
    valid_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(rows):
        errors = _validate_row(row, idx, folder)
        if errors:
            validation_errors.append({"index": idx + 2, "row": row, "errors": errors})
        else:
            valid_rows.append(row)

    # 自动分配时间为空的行
    if valid_rows:
        rows_with_empty_time = [r for r in valid_rows if not (r.get("scheduled_at") or "").strip()]
        if rows_with_empty_time and not args.date:
            raise SystemExit("存在未填写时间的行，请使用 --date 指定日期以自动分配时间。")
        if rows_with_empty_time and args.date:
            valid_rows = auto_assign_schedule(valid_rows, parse_date_for_filename(args.date))

    # 图文：校验达人须 TikTok Shop + PHOTO_SHOPPABLE_PERMISSION_PRODUCT
    row_index_by_id = {id(row): idx + 2 for idx, row in enumerate(rows)}
    photo_rows = [r for r in valid_rows if _row_media_type(r) == "photo"]
    if photo_rows:
        require_api_key()
        photo_usernames = list({r["creator_username"].strip() for r in photo_rows if r.get("creator_username")})
        photo_creator_map, _ = resolve_creator_batch(
            photo_usernames,
            platform="tiktok_shop",
            has_photo_permission=True,
        )
        still_valid: list[dict[str, Any]] = []
        for row in valid_rows:
            if _row_media_type(row) != "photo":
                still_valid.append(row)
                continue
            username = (row.get("creator_username") or "").strip()
            info = photo_creator_map.get(username.lower())
            excel_index = row_index_by_id.get(id(row), 0)
            if not info:
                validation_errors.append({
                    "index": excel_index,
                    "row": row,
                    "errors": [
                        f"未找到可发图文的 TikTok Shop 达人：{username}"
                        f"（需 authSource=TikTok Shop 且 hasPhotoPermission / {PHOTO_SHOPPABLE_PERMISSION}）"
                    ],
                })
                continue
            ok, reason = creator_can_publish_photo(info)
            if not ok:
                validation_errors.append({
                    "index": excel_index,
                    "row": row,
                    "errors": [reason],
                })
                continue
            still_valid.append(row)
        valid_rows = still_valid

    # dry-run
    if not args.confirm:
        results = _build_validation_results(rows, valid_rows, validation_errors, mode="dry-run")

        if args.format == "json":
            print_json(results)
        else:
            print(format_publish_results(results))
            print("\n这是 dry-run。加 --confirm 执行实际发布。", file=sys.stderr)
        return

    if validation_errors:
        results = _build_validation_results(rows, valid_rows, validation_errors, mode="needs-edit")
        if args.format == "json":
            print_json(results)
        else:
            print(format_publish_results(results))
        return

    # 批量查 creatorId（图文优先按 TikTok Shop 源解析）
    usernames = list({r["creator_username"].strip() for r in valid_rows})
    creator_map, not_found = resolve_creator_batch(usernames)
    photo_usernames = list({
        r["creator_username"].strip()
        for r in valid_rows
        if _row_media_type(r) == "photo"
    })
    if photo_usernames:
        shop_map, _ = resolve_creator_batch(
            photo_usernames,
            platform="tiktok_shop",
            has_photo_permission=True,
        )
        creator_map.update(shop_map)

    results: dict[str, Any] = {
        "mode": "live",
        "total": len(valid_rows),
        "succeeded": 0,
        "failed_count": 0,
        "video_type": _video_type_from_rows(valid_rows),
        "records_url": PUBLISH_RECORDS_URL,
        "rows": [],
    }

    for idx, row in enumerate(valid_rows):
        if idx > 0 and args.sleep_ms > 0:
            time.sleep(args.sleep_ms / 1000.0)

        username = row["creator_username"].strip()
        media_type = _row_media_type(row)
        media_names = _row_media_filenames(row)
        entry: dict[str, Any] = {
            "index": idx + 2,
            "creator_username": username,
            "platform": row.get("platform", ""),
            "media_type": media_type,
            "title": row.get("title", ""),
            "video_file": row.get("video_file", ""),
            "image_files": row.get("image_files", ""),
            "scheduled_at": row.get("scheduled_at", ""),
            "status": "failed",
            "post_id": None,
            "video_url": None,
            "post_status": None,
            "errors": [],
        }

        try:
            creator_info = creator_map.get(username.lower())
            if not creator_info:
                raise SystemExit(f"未找到创作者账号：{username}。请确认该账号已授权。")

            if media_type == "photo":
                ok, reason = creator_can_publish_photo(creator_info)
                if not ok:
                    raise SystemExit(reason)

            # 上传媒体（视频 1 个 / 图文多张）。
            # 图文：严格按 image_files 用户填写顺序上传；
            # businessId=首图 fileId，businessIds=全部 fileId（同序）。
            file_ids: list[str] = []
            last_url = ""
            for media_name in media_names:
                media_path = _resolve_media_path(folder, media_name)
                print(f"[{idx + 2}/{len(valid_rows)}] 上传：{media_name}", file=sys.stderr)
                default_ct = "image/jpeg" if media_type == "photo" else "video/mp4"
                upload_result = upload_file(str(media_path), default_content_type=default_ct)
                file_id = str(upload_result.get("id", "") or "")
                if not file_id:
                    raise SystemExit(f"上传成功但未返回 fileId：{media_name}")
                file_ids.append(file_id)
                last_url = upload_result.get("url", "") or last_url
            entry["video_url"] = last_url

            # 构建发布请求
            platform = row.get("platform", "tiktok_shop")
            tz_str = parse_timezone(row.get("timezone"))
            scheduled_at = None
            scheduled_tz = None
            if row.get("scheduled_at", "").strip():
                scheduled_at = parse_datetime_to_epoch_ms(row["scheduled_at"].strip(), tz_str)
                scheduled_tz = tz_str

            print(f"[{idx + 2}/{len(valid_rows)}] 发布：{username} - {row.get('title', '')}", file=sys.stderr)
            clean_title = sanitize_post_title(row.get("title", "") or "Untitled")
            raw_product_title = row.get("product_title") or ""
            clean_product_title = sanitize_product_title(raw_product_title) if raw_product_title else ""
            music_info = _row_music_info(row)

            post_result = create_post(
                creator_id=creator_info["creatorId"],
                title=clean_title or "Untitled",
                business_id=file_ids[0],
                platform=platform,
                scheduled_at=scheduled_at,
                scheduled_tz=scheduled_tz,
                oauth_region=creator_info.get("oauthRegion"),
                product_id=row.get("product_id") or None,
                product_title=clean_product_title,
                product_source=row.get("product_source") or "SHOP",
                media_type="PHOTO" if media_type == "photo" else "VIDEO",
                business_ids=file_ids if media_type == "photo" else None,
                photo_post_type=DEFAULT_PHOTO_POST_TYPE,
                music_info=music_info,
            )

            entry["status"] = "success"
            entry["post_id"] = post_result.get("postId") or str(post_result.get("id", ""))
            entry["post_status"] = post_result.get("status", "")
            entry["video_url"] = post_result.get("videoUrl") or entry["video_url"]

        except SystemExit as exc:
            entry["errors"].append(str(exc))
        except Exception as exc:
            entry["errors"].append(f"{type(exc).__name__}: {exc}")

        results["rows"].append(entry)
        if entry["status"] == "success":
            results["succeeded"] += 1
        else:
            results["failed_count"] += 1

    if args.format == "json":
        print_json(results)
    else:
        print(format_publish_results(results))
    if results["failed_count"] > 0:
        sys.exit(1)


def _build_validation_results(
    rows: list[dict[str, Any]],
    valid_rows: list[dict[str, Any]],
    validation_errors: list[dict[str, Any]],
    *,
    mode: str,
) -> dict[str, Any]:
    results: dict[str, Any] = {
            "mode": "dry-run",
            "total": len(rows),
            "dry_run_valid": len(valid_rows),
            "dry_run_invalid": len(validation_errors),
            "needs_edit_count": len(validation_errors),
            "succeeded": 0,
            "failed_count": 0,
            "rows": [],
            "validation_errors": validation_errors,
        }
    if mode == "needs-edit":
        results["mode"] = "needs-edit"

    for idx, row in enumerate(valid_rows):
        results["rows"].append({
            "index": idx + 2,
            "creator_username": row.get("creator_username", ""),
            "platform": row.get("platform", ""),
            "media_type": _row_media_type(row),
            "title": row.get("title", ""),
            "video_file": row.get("video_file", ""),
            "image_files": row.get("image_files", ""),
            "scheduled_at": row.get("scheduled_at", ""),
            "status": "dry-run",
        })
    for ve in validation_errors:
        results["rows"].append({
            "index": ve["index"],
            "creator_username": ve["row"].get("creator_username", ""),
            "platform": ve["row"].get("platform", ""),
            "media_type": _row_media_type(ve["row"]),
            "title": ve["row"].get("title", ""),
            "video_file": ve["row"].get("video_file", ""),
            "image_files": ve["row"].get("image_files", ""),
            "scheduled_at": ve["row"].get("scheduled_at", ""),
            "status": "needs-edit",
            "errors": ve["errors"],
        })
    return results


def _item_can_publish_photo(item: dict[str, Any]) -> bool:
    """pageList 原始 item 是否可发带货图文。"""
    summary = {
        "username": item.get("username") or "",
        "authSource": item.get("authSource") or item.get("auth_source") or "",
        "permissions": extract_permissions(item),
    }
    ok, _ = creator_can_publish_photo(summary)
    return ok


def _row_media_type(row: dict[str, Any]) -> str:
    """推断行媒体类型：显式 media_type，或根据 image_files / 文件扩展名。"""
    raw = (row.get("media_type") or "").strip()
    if raw:
        try:
            return parse_media_type(raw)
        except SystemExit:
            return "video"
    image_files = (row.get("image_files") or "").strip()
    if image_files:
        return "photo"
    video_file = (row.get("video_file") or "").strip()
    names = split_media_filenames(video_file)
    if names and all(Path(n).suffix.lower() in IMAGE_EXTENSIONS for n in names):
        return "photo"
    return "video"


def _row_media_filenames(row: dict[str, Any]) -> list[str]:
    """取出该行需要上传的媒体文件名列表。"""
    media_type = _row_media_type(row)
    if media_type == "photo":
        names = split_media_filenames(row.get("image_files"))
        if not names:
            names = split_media_filenames(row.get("video_file"))
        return names
    return split_media_filenames(row.get("video_file"))


def _resolve_media_path(folder: Path, name: str) -> Path:
    path = folder / name
    if path.exists():
        return path
    candidates = list(folder.glob(name))
    if candidates:
        return candidates[0]
    raise SystemExit(f"媒体文件不存在：{name}")


def _row_music_info(row: dict[str, Any]) -> dict[str, str] | None:
    music_id = (row.get("music_id") or "").strip()
    if not music_id:
        return None
    return {
        "id": music_id,
        "title": (row.get("music_title") or "").strip(),
        "author": (row.get("music_author") or "").strip(),
        "duration": (row.get("music_duration") or "").strip(),
    }


def _validate_row(row: dict[str, str], idx: int, folder: Path) -> list[str]:
    """逐行校验，返回错误信息列表。"""
    errors: list[str] = []

    creator = (row.get("creator_username") or "").strip()
    if not creator:
        errors.append("creator_username 为空")

    platform = (row.get("platform") or "").strip()
    if platform not in SUPPORTED_PLATFORMS:
        errors.append(f"platform 无效：{platform}（可选：{', '.join(SUPPORTED_PLATFORMS)}）")

    raw_media = (row.get("media_type") or "").strip()
    if raw_media:
        try:
            media_type = parse_media_type(raw_media)
        except SystemExit as exc:
            errors.append(str(exc))
            media_type = "video"
    else:
        media_type = _row_media_type(row)

    if media_type == "photo" and platform and platform != "tiktok_shop":
        errors.append("带货图文仅支持 platform=tiktok_shop")

    title = (row.get("title") or "").strip()
    if not title:
        errors.append("title 为空")
    elif has_unsupported_plain_text(title, POST_TITLE_MAX_LENGTH, allowed_symbols="#"):
        errors.append(f"title 不能超过 {POST_TITLE_MAX_LENGTH} 字符，且不能包含表情、标点或特殊符号（# 可用于 hashtag）")

    if platform == "tiktok_shop":
        product_id = (row.get("product_id") or "").strip()
        if not product_id:
            errors.append("带货 (tiktok_shop) 必须填写 产品ID")
        product_title = (row.get("product_title") or "").strip()
        if not product_title:
            errors.append("带货 (tiktok_shop) 必须填写 购物车标题")
        elif has_unsupported_plain_text(product_title, PRODUCT_TITLE_MAX_LENGTH):
            errors.append(f"购物车标题不能超过 {PRODUCT_TITLE_MAX_LENGTH} 字符，且不能包含表情、标点或特殊符号")

    scheduled_at = (row.get("scheduled_at") or "").strip()
    if scheduled_at:
        import re
        if not re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$", scheduled_at):
            errors.append(f"scheduled_at 格式错误：{scheduled_at}（需要 YYYY-MM-DD HH:MM）")

    tz_raw = (row.get("timezone") or "").strip()
    if tz_raw:
        try:
            parse_timezone(tz_raw)
        except SystemExit as exc:
            errors.append(str(exc))
    elif scheduled_at:
        errors.append("scheduled_at 已填写时，timezone 不能为空")

    media_names = _row_media_filenames(row)
    if not media_names:
        if media_type == "photo":
            errors.append(
                f"image_files 为空（至少 1 张、最多 {PHOTO_MAX_IMAGES} 张，多图用逗号分隔，如 a.jpg,b.jpg）"
            )
        else:
            errors.append("video_file 为空")
    elif media_type == "photo":
        resolved: list[Path] = []
        missing = False
        for name in media_names:
            try:
                resolved.append(_resolve_media_path(folder, name))
            except SystemExit:
                errors.append(f"图片文件不存在：{name}")
                missing = True
        if not missing:
            errors.extend(validate_photo_files(resolved))
    else:
        for name in media_names:
            try:
                _resolve_media_path(folder, name)
            except SystemExit:
                errors.append(f"视频文件不存在：{name}")

    music_id = (row.get("music_id") or "").strip()
    if music_id and media_type != "photo":
        errors.append("music_id 仅用于带货图文 (media_type=photo)")

    return errors


def _video_type_from_rows(rows: list[dict[str, Any]]) -> str:
    media_types = {_row_media_type(row) for row in rows}
    platforms = {row.get("platform") for row in rows if row.get("platform")}
    if media_types == {"photo"} and platforms == {"tiktok_shop"}:
        return "带货图文"
    if platforms == {"tiktok_shop"}:
        return "带货"
    if platforms == {"tiktok"}:
        return "普通/养号"
    return "混合"


# ── fill ─────────────────────────────────────────────────────

def command_fill(args: argparse.Namespace) -> None:
    """更新 CSV/XLSX 排期表中的单元格。"""
    folder = Path(args.folder).expanduser()
    schedule_file = folder / "schedule.csv"
    if not schedule_file.exists():
        schedule_file = folder / "schedule.xlsx"
    if not schedule_file.exists():
        raise SystemExit(f"文件夹内未找到排期表：{folder}")

    if schedule_file.suffix.lower() == ".csv":
        command_fill_csv(args, schedule_file)
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("需要 openpyxl 库：pip install openpyxl")

    wb = load_workbook(str(schedule_file))
    ws = wb.active
    if ws is None:
        raise SystemExit("无法打开工作表。")

    # 读取表头映射
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        eng = COLUMN_LABEL_TO_KEY.get(h, h)
        col_map[eng] = idx

    target_col = COLUMN_LABEL_TO_KEY.get(args.col, args.col)
    if target_col not in col_map:
        avail = [COLUMN_LABELS.get(c, c) for c in col_map]
        raise SystemExit(f"找不到列「{args.col}」，可用列：{', '.join(avail)}")
    col_idx = col_map[target_col]

    updated = 0

    # 自动填产品标题
    if args.auto_product_title:
        require_api_key()
        pid_col_idx = col_map.get("product_id", -1)
        updated_count = 0
        for row_idx in range(2, ws.max_row + 1):
            pid = str(ws.cell(row=row_idx, column=pid_col_idx + 1).value or "").strip()
            if not pid:
                continue
            creator = str(ws.cell(row=row_idx, column=col_map.get("creator_username", 0) + 1).value or "").strip()
            if not creator:
                continue
            try:
                # 尝试从 shop 产品搜索获取标题
                from .api import list_shop_products, resolve_creator_batch
                found, _ = resolve_creator_batch([creator])
                cinfo = found.get(creator.lower(), {})
                cid = (cinfo or {}).get("creatorId", "")
                if cid:
                    result = list_shop_products(creator_id=cid, page_size=5, title_keyword=pid)
                    products = (result.get("data") or {}).get("products") or []
                    # 匹配 product id
                    for p in products:
                        if str(p.get("id", "")) == pid:
                            from .excel_utils import sanitize_product_title
                            title = sanitize_product_title(p.get("title", ""))
                            ws.cell(row=row_idx, column=col_idx + 1).value = title
                            updated_count += 1
                            break
            except SystemExit:
                pass  # 单个失败跳过，继续其他行
        updated = updated_count
        print(f"已自动填入 {updated} 个购物车标题", file=sys.stderr)

    # 按行号填充
    elif args.row is not None:
        row_idx = args.row + 2  # 0-indexed → Excel data row (header=1)
        ws.cell(row=row_idx, column=col_idx + 1).value = args.value
        updated = 1

    # 按 creator 筛选填充
    elif args.creator:
        creator_col_idx = col_map.get("creator_username", 0)
        for row_idx in range(2, ws.max_row + 1):
            c = str(ws.cell(row=row_idx, column=creator_col_idx + 1).value or "").strip()
            if c.lower() == args.creator.lower():
                ws.cell(row=row_idx, column=col_idx + 1).value = args.value
                updated += 1

    # 填充所有行
    elif args.value:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx + 1).value = args.value
            updated = ws.max_row - 1

    wb.save(str(schedule_file))
    wb.close()

    if args.format == "json":
        print_json({"updated": updated, "col": target_col})
    else:
        print(f"已更新 {updated} 行：{COLUMN_LABELS.get(target_col, target_col)}")


def command_fill_csv(args: argparse.Namespace, schedule_file: Path) -> None:
    """更新 CSV 排期表。"""
    current_columns = read_csv_columns(str(schedule_file))
    rows = parse_excel_or_csv(str(schedule_file))
    target_col = COLUMN_LABEL_TO_KEY.get(args.col, args.col)
    if target_col not in current_columns:
        avail = [COLUMN_LABELS.get(c, c) for c in current_columns]
        raise SystemExit(f"找不到列「{args.col}」，可用列：{', '.join(avail)}")

    updated = 0
    if args.auto_product_title:
        raise SystemExit("CSV 模式暂不支持 --auto-product-title，请直接填 product_title 或改用具体 --value。")
    if args.row is not None:
        if args.row < 0 or args.row >= len(rows):
            raise SystemExit(f"行号超出范围：{args.row}（当前 {len(rows)} 行数据）")
        rows[args.row][target_col] = args.value
        updated = 1
    elif args.creator:
        for row in rows:
            if (row.get("creator_username") or "").strip().lower() == args.creator.lower():
                row[target_col] = args.value
                updated += 1
    elif args.value:
        for row in rows:
            row[target_col] = args.value
        updated = len(rows)

    write_csv_schedule(str(schedule_file), rows, columns=current_columns)

    if args.format == "json":
        print_json({"updated": updated, "col": target_col})
    else:
        print(f"已更新 {updated} 行：{COLUMN_LABELS.get(target_col, target_col)}")


# ── products search ──────────────────────────────────────────

def command_products_search(args: argparse.Namespace) -> None:
    username = args.creator_username.strip()
    platform = args.platform or "tiktok_shop"
    source = args.source or "shop"

    require_api_key()

    creator_info = resolve_creator_batch([username], platform=platform)[0].get(username.lower())
    if not creator_info:
        raise SystemExit(f"未找到创作者账号：{username}")

    creator_id = creator_info["creatorId"]
    if not creator_id:
        raise SystemExit(f"创作者 {username} 缺少 creatorId")

    if source == "shop":
        result = list_shop_products(
            creator_id=creator_id,
            page_size=args.page_size,
            title_keyword=args.keyword or None,
        )
    else:
        from .api import list_showcase_products
        result = list_showcase_products(
            creator_id=creator_id,
            page_size=args.page_size,
        )

    data = result.get("data") or {}
    products = data.get("products") or []

    if args.keyword and source == "shop":
        keyword = args.keyword.strip().lower()
        products = [p for p in products if keyword in (p.get("title") or "").lower()]

    if args.format == "json":
        print_json({"products": products, "total": len(products)})
        return

    print(format_products(products, source))


# ── argparse helpers ─────────────────────────────────────────

def add_format_argument(parser: argparse.ArgumentParser, default: str = "json") -> None:
    parser.add_argument("--format", choices=("json", "text"), default=default, help="输出格式。")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="灵途批量视频发布。")
    subparsers = parser.add_subparsers(dest="command", required=True)

    # gen-csv
    p = subparsers.add_parser("gen-csv", help="生成排期 CSV 模板（桌面文件夹）。")
    p.add_argument("--platform", choices=SUPPORTED_PLATFORMS, required=True, help="发布平台。")
    p.add_argument(
        "--media-type",
        choices=SUPPORTED_MEDIA_TYPES,
        default="video",
        help="媒体类型：video=视频（默认），photo=带货图文（仅 tiktok_shop）。",
    )
    p.add_argument("--creators", default=None, help="逗号分隔的达人用户名，不传则取全部已授权。")
    p.add_argument("--date", required=True, help="起始日期 YYYY-MM-DD。")
    p.add_argument("--days", type=int, default=1, help="连续发布天数，默认 1。")
    p.add_argument("--count", type=int, default=3, help="每达人每日发布条数，默认 3。")
    p.add_argument("--daily-counts", default="", help="按日期指定每达人发布条数，如 2026-07-06=2,2026-07-07=3。传入后优先使用这些日期。")
    p.add_argument("--product-id", default="", help="产品 ID（仅 tiktok_shop）。")
    p.add_argument("--timezone", default="", help="时区简码或 IANA，如 EST、PST、CN、America/New_York。不传则按达人区域自动推断。")
    p.add_argument("--region", "--country", dest="region", default="", help="目标地区/国家，如 US。tiktok_shop 列表可按地区筛选；tiktok 普通视频仅用于默认时区。")
    p.add_argument("--output-dir", default=None, help="自定义输出目录，默认 ~/Desktop/视频发布_{date}/ 或 图文发布_{date}/。")
    p.add_argument("--dry-run", action="store_true", help="仅预览排期表，不生成文件。")
    add_format_argument(p, default="text")
    p.set_defaults(func=command_gen_csv)

    # creators
    p = subparsers.add_parser("creators", help="列出已授权的创作者账号。")
    p.add_argument("--platform", choices=SUPPORTED_PLATFORMS, default=None, help="筛选平台。")
    p.add_argument("--username", default="", help="按用户名搜索。")
    p.add_argument("--region", "--country", dest="region", default="", help="地区筛选，仅 tiktok_shop 带货达人列表支持。")
    p.add_argument(
        "--has-photo-permission",
        action="store_true",
        help="仅拉取有带货图文权限的账号（查询参数 hasPhotoPermission=true）。",
    )
    p.add_argument("--page-size", type=int, default=200, help="每页数量。")
    add_format_argument(p, default="text")
    p.set_defaults(func=command_creators)

    # publish
    p = subparsers.add_parser("publish", help="从文件夹读取排期表并执行发布。")
    p.add_argument("--folder", required=True, help="排期文件夹路径。")
    p.add_argument("--date", default="", help="自动排期日期 YYYY-MM-DD（未填时间的行必传）。")
    p.add_argument("--confirm", action="store_true", help="确认发布（默认 dry-run）。")
    p.add_argument("--sleep-ms", type=int, default=500, help="每行操作间隔毫秒，默认 500。")
    add_format_argument(p, default="text")
    p.set_defaults(func=command_publish)

    # fill
    p = subparsers.add_parser("fill", help="更新 CSV/XLSX 排期表单元格。")
    p.add_argument("--folder", required=True, help="排期文件夹路径。")
    p.add_argument("--col", required=True, help="要填充的列名（英文 key 或中文名）。")
    p.add_argument("--value", default="", help="填充的值（所有行/按 --creator/按 --row）。")
    p.add_argument("--row", type=int, default=None, help="行号（0=第一行数据）。")
    p.add_argument("--creator", default=None, help="只填充指定达人的行。")
    p.add_argument("--auto-product-title", action="store_true", help="自动从产品 API 搜索并填入购物车标题。")
    add_format_argument(p, default="text")
    p.set_defaults(func=command_fill)

    # products
    pp = subparsers.add_parser("products", help="商品查询（子命令）。")
    pps = pp.add_subparsers(dest="products_command", required=True)
    ps = pps.add_parser("search", help="搜索创作者的商品。")
    ps.add_argument("--creator-username", required=True, help="创作者用户名。")
    ps.add_argument("--platform", choices=SUPPORTED_PLATFORMS, default="tiktok_shop", help="平台。")
    ps.add_argument("--source", choices=("shop", "showcase"), default="shop", help="商品来源。")
    ps.add_argument("--keyword", default="", help="商品标题关键词。")
    ps.add_argument("--page-size", type=int, default=20, help="每页数量。")
    add_format_argument(ps, default="text")
    ps.set_defaults(func=command_products_search)

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


def run() -> None:
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
