"""CSV 模板生成、CSV/XLSX 通用读取。"""

from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path
from typing import Any

from .config import (
    COLUMN_LABELS,
    COLUMN_LABEL_TO_KEY,
    CSV_ALL_COLUMNS,
    PRODUCT_SOURCE_OPTIONS,
    PRODUCT_TITLE_MAX_LENGTH,
    POST_TITLE_MAX_LENGTH,
    SUPPORTED_PLATFORMS,
    TIMEZONE_DROPDOWN_OPTIONS,
)


def parse_excel_or_csv(path: str) -> list[dict[str, str]]:
    """读取 .xlsx 或 .csv 文件，自动映射中/英文列名，返回英 key 行字典列表。"""
    p = Path(path).expanduser()
    if not p.exists():
        raise SystemExit(f"排期文件不存在：{p}")

    suffix = p.suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(p)
    elif suffix in (".xlsx", ".xls"):
        rows = _parse_xlsx(p)
    else:
        raise SystemExit(f"不支持的文件格式：{suffix}，仅支持 .csv / .xlsx")

    return [_normalize_keys(row) for row in rows]


def _normalize_keys(row: dict[str, str]) -> dict[str, str]:
    """将中文列名映射为英文 key。"""
    normalized: dict[str, str] = {}
    for key, value in row.items():
        k = key.strip()
        if k in COLUMN_LABEL_TO_KEY:
            normalized[COLUMN_LABEL_TO_KEY[k]] = value
        elif k in COLUMN_LABELS.values():
            # Already a known Chinese label
            normalized[COLUMN_LABEL_TO_KEY[k]] = value
        else:
            normalized[k] = value
    return normalized


def _parse_csv(file_path: Path) -> list[dict[str, str]]:
    """解析 CSV 文件（utf-8-sig，BOM 兼容）。"""
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise SystemExit("CSV 文件为空或缺少表头。")
        rows: list[dict[str, str]] = []
        for row in reader:
            normalized: dict[str, str] = {}
            for key, value in row.items():
                normalized[key.strip()] = (value or "").strip()
            if any(v for v in normalized.values()):
                rows.append(normalized)
        return rows


def _parse_xlsx(file_path: Path) -> list[dict[str, str]]:
    """解析 Excel 文件。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("读取 Excel 需要 openpyxl 库：pip install openpyxl")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws is None:
        raise SystemExit("Excel 文件没有工作表。")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        raise SystemExit("Excel 文件为空。")

    headers = [str(h or "").strip() for h in headers_raw]
    if not headers:
        raise SystemExit("Excel 文件缺少表头。")

    rows: list[dict[str, str]] = []
    for row_values in rows_iter:
        row_dict: dict[str, str] = {}
        for idx, value in enumerate(row_values):
            if idx < len(headers):
                col = headers[idx]
                row_dict[col] = str(value).strip() if value is not None else ""
        if any(v for v in row_dict.values()):
            rows.append(row_dict)

    wb.close()
    return rows


def generate_csv_template(
    output_path: str,
    rows_data: list[dict[str, str]],
) -> str:
    """生成 CSV 排期表，返回输出路径。"""
    out = Path(output_path).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    headers = [COLUMN_LABELS.get(c, c) for c in CSV_ALL_COLUMNS]
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows_data:
            writer.writerow({COLUMN_LABELS.get(key, key): row.get(key, "") for key in CSV_ALL_COLUMNS})
    return str(out)


def write_csv_schedule(output_path: str, rows_data: list[dict[str, str]]) -> str:
    """把英文 key 行数据写回 CSV 排期表。"""
    return generate_csv_template(output_path, rows_data)


def generate_excel_template(
    output_path: str,
    creators: list[str] | None = None,
    platform: str = "tiktok_shop",
    date: str = "",
    count: int = 3,
    times: list[str] | None = None,
    timezone: str = "",
    product_id: str = "",
    product_source: str = "SHOP",
    rows_data: list[dict[str, str]] | None = None,
) -> str:
    """生成带下拉校验的 Excel 排期模板，返回输出路径。

    可通过 rows_data 直接传入预构建的行列表，此时忽略 creators/date/count/times 参数。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise SystemExit("生成 Excel 需要 openpyxl 库：pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise SystemExit("无法创建工作表。")
    ws.title = "排期表"

    headers = [COLUMN_LABELS.get(c, c) for c in CSV_ALL_COLUMNS]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    if rows_data is None:
        times = times or ["09:00", "14:00", "19:00"]
        rows_data = []
        for creator in (creators or []):
            for t in times:
                rows_data.append({
                    "creator_username": creator,
                    "platform": platform,
                    "title": "",
                    "product_id": product_id if platform == "tiktok_shop" else "",
                    "product_source": product_source if platform == "tiktok_shop" else "",
                    "scheduled_at": f"{date} {t}",
                    "timezone": timezone,
                    "video_file": "",
                })

    for row_idx, row_data in enumerate(rows_data, 2):
        for col_idx, eng_key in enumerate(CSV_ALL_COLUMNS, 1):
            value = row_data.get(eng_key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if eng_key == "scheduled_at":
                cell.alignment = Alignment(horizontal="center")
            elif eng_key in ("title", "product_title"):
                cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 30

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions or "A1:I1"

    out = Path(output_path).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    wb.close()
    return str(out)


def _col_letter(idx: int) -> str:
    """列索引(1-based) → Excel 列字母。"""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(rem + 65) + letters
    return letters


def parse_timezone(raw: str | None) -> str:
    """解析时区：短码 → IANA，验证 IANA 有效性，失败则报错。"""
    from .config import TIMEZONE_MAP

    if not raw or not raw.strip():
        return ""

    value = raw.strip()

    # 解析下拉格式 "America/New_York (EST)" → "America/New_York"
    match = re.match(r"^([A-Za-z/_]+)\s*\(.*\)", value)
    if match:
        value = match.group(1).strip()

    if value in TIMEZONE_MAP:
        return TIMEZONE_MAP[value]

    try:
        from zoneinfo import ZoneInfo
        ZoneInfo(value)
        return value
    except Exception:
        raise SystemExit(
            f"无效时区：{raw}。请使用 IANA 时区（如 America/New_York）"
            f"或短码（{', '.join(sorted(TIMEZONE_MAP.keys()))}）。"
        )


def parse_datetime_to_epoch_ms(scheduled_at: str, tz_iana: str) -> int:
    """'YYYY-MM-DD HH:MM' + IANA 时区 → 毫秒时间戳。"""
    from datetime import datetime
    from zoneinfo import ZoneInfo

    try:
        dt = datetime.strptime(scheduled_at.strip(), "%Y-%m-%d %H:%M")
        tz = ZoneInfo(tz_iana)
        localized = dt.replace(tzinfo=tz)
        return int(localized.timestamp() * 1000)
    except ValueError as exc:
        raise SystemExit(f"时间格式错误：{scheduled_at}（需要 YYYY-MM-DD HH:MM）") from exc


def sanitize_plain_text(value: str, max_length: int, allowed_symbols: str = "") -> str:
    """清理平台不支持的表情、符号、标点，并截断到最大长度。"""
    normalized = unicodedata.normalize("NFKC", value or "")
    chars: list[str] = []
    previous_space = False
    for char in normalized:
        category = unicodedata.category(char)
        if char in allowed_symbols:
            chars.append(char)
            previous_space = False
        elif category[0] in ("L", "N") or category == "Zs" or char in ("\n", "\r", "\t"):
            if char in ("\n", "\r", "\t") or category == "Zs":
                if not previous_space:
                    chars.append(" ")
                    previous_space = True
            else:
                chars.append(char)
                previous_space = False
    return "".join(chars).strip()[:max_length]


def has_unsupported_plain_text(value: str, max_length: int, allowed_symbols: str = "") -> bool:
    """判断文本是否超长或包含平台不支持的表情/符号/标点。"""
    raw = (value or "").strip()
    return len(raw) > max_length or sanitize_plain_text(raw, max_length, allowed_symbols) != raw


def sanitize_product_title(title: str, max_length: int = PRODUCT_TITLE_MAX_LENGTH) -> str:
    """清理购物车标题：去表情和特殊符号，截断至 max_length 字符。"""
    return sanitize_plain_text(title, max_length)


def sanitize_post_title(title: str, max_length: int = POST_TITLE_MAX_LENGTH) -> str:
    """清理视频文案：去表情和特殊符号，保留 #，截断至 max_length 字符。"""
    return sanitize_plain_text(title, max_length, allowed_symbols="#")


def parse_date_for_filename(date_str: str) -> str:
    """验证并标准化日期格式 YYYY-MM-DD。"""
    import re as _re
    if not _re.match(r"^\d{4}-\d{2}-\d{2}$", date_str.strip()):
        raise SystemExit(f"日期格式错误：{date_str}（需要 YYYY-MM-DD）")
    return date_str.strip()
